#!/usr/bin/env python3
"""
1029 이태원 참사 언론 기사 수집
- Google News RSS (키 불필요)
- Naver News API (NAVER_CLIENT_ID / NAVER_CLIENT_SECRET)

BigKinds: 후순위 · 미구현

Usage:
  python3 scripts/1029_collect.py
  python3 scripts/1029_collect.py --source google|naver|all
  python3 scripts/1029_collect.py --export-xlsx
  python3 scripts/1029_collect.py --import-seed /path/to.xlsx
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from pathlib import Path
from typing import Any, Optional
from urllib.parse import parse_qs, urlencode, urlparse, urlunparse

import feedparser
import requests
import yaml

BASE = Path(__file__).resolve().parent.parent
DATA_PATH = BASE / "data" / "1029-articles.json"
QUERIES_PATH = Path(__file__).resolve().parent / "1029_queries.yaml"
XLSX_PATH = BASE / "data" / "1029-articles.xlsx"

NAVER_CLIENT_ID = os.getenv("NAVER_CLIENT_ID", "")
NAVER_CLIENT_SECRET = os.getenv("NAVER_CLIENT_SECRET", "")

STRIP_HTML = re.compile(r"<[^>]+>")
TITLE_PUB = re.compile(r"\s+[-\u2013\u2014]\s+([^\-\u2013\u2014]+)\s*$")


def today() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d")


def load_queries() -> dict:
    with open(QUERIES_PATH, encoding="utf-8") as f:
        return yaml.safe_load(f)


def canon_url(url: str) -> str:
    if not url:
        return ""
    p = urlparse(url.strip())
    # drop common tracking params
    q = parse_qs(p.query, keep_blank_values=False)
    for k in list(q.keys()):
        if k.lower() in {"input", "from", "utm_source", "utm_medium", "utm_campaign", "sid"}:
            del q[k]
    query = urlencode({k: v[0] if len(v) == 1 else v for k, v in q.items()}, doseq=True)
    path = p.path.rstrip("/") or "/"
    return urlunparse((p.scheme, p.netloc.lower(), path, "", query, ""))


def item_id(url: str, published: str) -> str:
    h = hashlib.md5(f"{canon_url(url)}|{published}".encode()).hexdigest()[:10]
    return f"a1029-{published}-{h}"


def empty_db(since: str = "2022-10-29") -> dict:
    return {
        "meta": {
            "title": "1029 이태원 참사 언론 목록",
            "since": since,
            "last_updated": None,
            "total_count": 0,
            "sources": ["google_news", "naver"],
        },
        "items": [],
    }


def load_db() -> dict:
    if DATA_PATH.exists():
        with open(DATA_PATH, encoding="utf-8") as f:
            return json.load(f)
    return empty_db()


def save_db(db: dict) -> None:
    db["meta"]["total_count"] = len(db["items"])
    db["meta"]["last_updated"] = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(DATA_PATH, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)
        f.write("\n")


def existing_canons(db: dict) -> set[str]:
    return {canon_url(i.get("url", "")) for i in db.get("items", []) if i.get("url")}


def split_title_publisher(title: str) -> tuple[str, str]:
    m = TITLE_PUB.search(title)
    if m:
        pub = m.group(1).strip()
        clean = TITLE_PUB.sub("", title).strip()
        return clean, pub
    return title.strip(), ""


def make_item(
    *,
    title: str,
    url: str,
    published_at: str,
    author: str = "",
    publisher: str = "",
    source: str,
    query_id: str,
    collector: str = "auto",
    note: str = "",
    collected_at: Optional[str] = None,
) -> dict[str, Any]:
    if not publisher:
        title, publisher = split_title_publisher(title)
    return {
        "id": item_id(url, published_at or today()),
        "title": STRIP_HTML.sub("", title).replace("&quot;", '"').replace("&amp;", "&").strip(),
        "published_at": published_at or today(),
        "author": author or "",
        "publisher": publisher or "",
        "url": url,
        "url_canon": canon_url(url),
        "note": note,  # auto rows stay ""
        "collected_at": collected_at or today(),
        "collector": collector,
        "source": source,
        "query_id": query_id,
    }


def google_rss_url(q: str) -> str:
    from urllib.parse import quote

    return (
        "https://news.google.com/rss/search?"
        f"q={quote(q)}&hl=ko&gl=KR&ceid=KR:ko"
    )


def collect_google(queries: list[dict], since: str) -> list[dict]:
    out: list[dict] = []
    for q in queries:
        qid = q["id"]
        gq = q.get("google") or q.get("naver")
        if not gq:
            continue
        url = google_rss_url(gq)
        print(f"[Google] {qid} …")
        try:
            feed = feedparser.parse(url)
            n = 0
            for entry in feed.entries:
                link = entry.get("link", "")
                title = entry.get("title", "")
                if not link or not title:
                    continue
                published = today()
                if entry.get("published_parsed"):
                    published = datetime(*entry.published_parsed[:6]).strftime("%Y-%m-%d")
                if published < since:
                    continue
                source_name = ""
                if entry.get("source") and entry.source.get("title"):
                    source_name = entry.source.title
                item = make_item(
                    title=title,
                    url=link,
                    published_at=published,
                    publisher=source_name,
                    source="google_news",
                    query_id=qid,
                )
                out.append(item)
                n += 1
            print(f"[Google] {qid}: {n} kept (feed {len(feed.entries)})")
        except Exception as e:
            print(f"[Google] {qid}: ERROR {e}")
    return out


def collect_naver(queries: list[dict], since: str) -> list[dict]:
    if not NAVER_CLIENT_ID or not NAVER_CLIENT_SECRET:
        print("[Naver] skip — set NAVER_CLIENT_ID / NAVER_CLIENT_SECRET")
        return []
    headers = {
        "X-Naver-Client-Id": NAVER_CLIENT_ID,
        "X-Naver-Client-Secret": NAVER_CLIENT_SECRET,
    }
    out: list[dict] = []
    for q in queries:
        qid = q["id"]
        nq = q.get("naver")
        if not nq:
            continue
        print(f"[Naver] {qid} '{nq}' …")
        try:
            resp = requests.get(
                "https://openapi.naver.com/v1/search/news.json",
                headers=headers,
                params={"query": nq, "display": 100, "sort": "date"},
                timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()
            n = 0
            for row in data.get("items", []):
                link = row.get("originallink") or row.get("link", "")
                title = STRIP_HTML.sub("", row.get("title", ""))
                if not link or not title:
                    continue
                published = today()
                pub = row.get("pubDate", "")
                if pub:
                    try:
                        published = parsedate_to_datetime(pub).strftime("%Y-%m-%d")
                    except Exception:
                        pass
                if published < since:
                    continue
                item = make_item(
                    title=title,
                    url=link,
                    published_at=published,
                    source="naver",
                    query_id=qid,
                )
                out.append(item)
                n += 1
            print(f"[Naver] {qid}: {n} kept")
        except Exception as e:
            print(f"[Naver] {qid}: ERROR {e}")
    return out


def merge(db: dict, new_items: list[dict]) -> int:
    seen = existing_canons(db)
    added = 0
    for it in new_items:
        c = it.get("url_canon") or canon_url(it.get("url", ""))
        if not c or c in seen:
            continue
        # auto: note always empty
        if it.get("collector") == "auto":
            it["note"] = ""
        db["items"].append(it)
        seen.add(c)
        added += 1
    db["items"].sort(key=lambda x: x.get("published_at") or "", reverse=True)
    return added


def export_xlsx(db: dict, path: Path = XLSX_PATH) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        print("openpyxl required for xlsx export")
        sys.exit(1)
    wb = Workbook()
    ws = wb.active
    ws.title = "전체"
    headers = ["기사제목", "발행날짜", "기자명", "발행처", "url", "비고", "작성일", "작성자"]
    ws.append(headers)
    for it in db.get("items", []):
        # 기자명: 빅카인즈면 author(기고자). Google 등은 비움
        # 작성자(기입자)·비고: 항상 비움
        reporter = it.get("author") or it.get("reporter") or ""
        if it.get("source") != "bigkinds":
            reporter = ""
        ws.append(
            [
                it.get("title", ""),
                it.get("published_at", ""),
                reporter,
                it.get("publisher", ""),
                it.get("url", ""),
                "",
                it.get("collected_at", ""),
                "",
            ]
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    print(f"[xlsx] wrote {path} ({len(db.get('items', []))} rows)")


def import_seed(path: Path, db: dict) -> int:
    """Seed xlsx columns: 0제목 1발행날짜 2기자 3발행처 4url 5비고 6작성일 7기입자"""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0
    added = 0
    seen = existing_canons(db)
    for row in rows[1:]:
        if not row or not row[0]:
            continue
        title = str(row[0]).strip()
        url = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        if not title or not url:
            continue
        pub = row[1] if len(row) > 1 else None
        if hasattr(pub, "strftime"):
            published = pub.strftime("%Y-%m-%d")
        else:
            published = str(pub)[:10] if pub else today()
        author = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        publisher = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        # 비고는 시드에 있어도 가져오지 않음 (담당자 전용)
        note = ""
        collected = row[6] if len(row) > 6 else None
        if hasattr(collected, "strftime"):
            collected_at = collected.strftime("%Y-%m-%d")
        else:
            collected_at = str(collected)[:10] if collected else today()
        collector = str(row[7]).strip() if len(row) > 7 and row[7] else "seed"
        c = canon_url(url)
        if c in seen:
            continue
        it = make_item(
            title=title,
            url=url,
            published_at=published,
            author=author,
            publisher=publisher,
            source="seed",
            query_id="seed",
            collector=collector,
            note=note,
            collected_at=collected_at,
        )
        db["items"].append(it)
        seen.add(c)
        added += 1
    db["items"].sort(key=lambda x: x.get("published_at") or "", reverse=True)
    return added


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", choices=["all", "google", "naver"], default="all")
    ap.add_argument("--export-xlsx", action="store_true")
    ap.add_argument("--import-seed", type=Path, default=None)
    args = ap.parse_args()

    cfg = load_queries()
    since = str(cfg.get("since") or "2022-10-29")
    queries = cfg.get("queries") or []
    db = load_db()
    db["meta"]["since"] = since

    if args.import_seed:
        n = import_seed(args.import_seed, db)
        save_db(db)
        print(f"[seed] imported {n} new rows → total {len(db['items'])}")
        if args.export_xlsx:
            export_xlsx(db)
        return

    new: list[dict] = []
    if args.source in ("all", "google"):
        new.extend(collect_google(queries, since))
    if args.source in ("all", "naver"):
        new.extend(collect_naver(queries, since))

    added = merge(db, new)
    save_db(db)
    print(f"[done] +{added} → total {len(db['items'])} → {DATA_PATH}")

    if args.export_xlsx or True:
        # always refresh xlsx for download link
        export_xlsx(db)


if __name__ == "__main__":
    main()
